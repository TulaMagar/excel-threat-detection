from django.http import response
from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm

from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .decorators import unauthenticated_user, allowed_users
from django.contrib.auth.models import Group
from openpyxl import load_workbook
import pandas as pd
import os
import shutil
from oletools.olevba3 import VBA_Parser
import re

@login_required(login_url='/registration/login/')
def login_request(request):
    if request.method == 'POST':
        form = AuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.info(request, f"You are now logged in as {username}")
            return render(request, "upload.html")

    form = AuthenticationForm()
    return render(request = request,
                    template_name = "login.html",
                    context={"form":form})

@unauthenticated_user
def login(request):
    """
    Creates login view
    Returns: rendered login page
    """
    return render(request, 'login.html')


def register(response):
    if response.method == 'POST':
        form = UserCreationForm(response.POST)
        if form.is_valid():
            user = form.save()           
            if response.POST.get('group') == 'current':
                group = Group.objects.get(name='New user')
            user.groups.add(group)
            return redirect("login")
    else:
        form = UserCreationForm()
    return render(response, 'registration/register.html', {'form': form})

def logout_request(request):
    logout(request)
    messages.info(request, "you have successfully logged out.")
    return redirect('login')


@allowed_users(allowed_roles=["creator"])
def upload(request):
    if request.method == 'POST':
        if request.FILES.get('document'):
            file = request.FILES['document']

            EXCEL_FILE_EXTENSIONS = ('xlsb', 'xls', 'xlsm', 'xla', 'xlt', 'xlam',)
            KEEP_NAME = False  # Set this to True if you would like to keep "Attribute VB_Name"

            result = []
            # macros check
            vba_parser = VBA_Parser(file)
            vba_modules = vba_parser.extract_all_macros()
            if vba_parser.detect_vba_macros():
                result.append("Caution, macros has been found in your excel file.")


            df = pd.read_excel(file, index_col=0)

            count = 0
            
            # check .exe, url 
            for ind, column in enumerate(df.columns):
                count = count + 1
                for value in df[column]:
                    if re.findall(r'\b\S*\.exe\b', str(value)):
                        message = 'found .exe file', str(count)
                        result.append(message)

                    elif re.findall('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', str(value)):
                        message = 'found a url link', str(count)
                        result.append(message)
            
            # check hidden rows
            workbook = load_workbook(filename=file, data_only=True)
            xls = workbook[workbook.sheetnames[0]] 

            hidden_rows = []
            for rowLetter,rowDimension in xls.row_dimensions.items():
                if rowDimension.hidden == True:
                    hidden_row = 'found hidden rows at ', rowLetter
                    hidden_rows.append(hidden_row)


            dic_result = { 'report': result}
            return render(request, 'upload.html', dic_result)
    return render(request, 'upload.html')